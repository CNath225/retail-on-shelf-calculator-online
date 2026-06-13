import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from alias_workbook import (
    ALIAS_SHEET_NAME,
    copy_workbook_with_alias_decisions,
    create_alias_decisions_workbook,
    load_alias_decisions_from_workbook,
    write_alias_sheet_xlsxwriter,
)


class AliasWorkbookTests(unittest.TestCase):
    def test_copy_workbook_embeds_hidden_alias_sheet(self):
        decisions = [
            {
                "domain": "sku",
                "raw_value": "Aqua10 Roller AE (JB)",
                "canonical": "Aqua10 Roller AE",
                "action": "map",
                "channel": "B",
            }
        ]
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "template.xlsx"
            target = tmp_path / "template_with_alias.xlsx"
            pd.DataFrame([{"Country": "AU", "Category": "Robot", "Channel": "JB", "SKU": "TTL"}]).to_excel(
                source,
                sheet_name="ANZ On-Shelf Retailer",
                index=False,
            )

            copy_workbook_with_alias_decisions(source, target, decisions)

            workbook = load_workbook(target)
            try:
                self.assertIn(ALIAS_SHEET_NAME, workbook.sheetnames)
                self.assertEqual(workbook[ALIAS_SHEET_NAME].sheet_state, "hidden")
                self.assertIn("ANZ On-Shelf Retailer", workbook.sheetnames)
            finally:
                workbook.close()
            self.assertEqual(load_alias_decisions_from_workbook(target), decisions)

    def test_xlsxwriter_hidden_alias_sheet_round_trips(self):
        decisions = [
            {
                "domain": "account",
                "raw_value": "Harvey Norman (HN)",
                "canonical": "HN",
                "action": "map",
            }
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"
            with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
                pd.DataFrame([{"Country": "AU"}]).to_excel(writer, sheet_name="Report Preview", index=False)
                write_alias_sheet_xlsxwriter(writer.book, decisions)

            workbook = load_workbook(path)
            try:
                self.assertEqual(workbook[ALIAS_SHEET_NAME].sheet_state, "hidden")
            finally:
                workbook.close()
            self.assertEqual(load_alias_decisions_from_workbook(path), decisions)

    def test_alias_carrier_workbook_has_visible_sheet_and_hidden_map(self):
        decisions = [
            {
                "domain": "country",
                "raw_value": "Australia",
                "canonical": "AU",
                "action": "map",
            }
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "alias_carrier.xlsx"
            create_alias_decisions_workbook(path, decisions)
            workbook = load_workbook(path)
            try:
                self.assertIn("Alias Carrier", workbook.sheetnames)
                self.assertEqual(workbook["Alias Carrier"].sheet_state, "visible")
                self.assertEqual(workbook[ALIAS_SHEET_NAME].sheet_state, "hidden")
            finally:
                workbook.close()
            self.assertEqual(load_alias_decisions_from_workbook(path), decisions)


if __name__ == "__main__":
    unittest.main()
