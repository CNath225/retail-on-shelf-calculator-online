import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from step4_generate_report_preview import (
    apply_presentation_formatting,
    apply_report_preview_formatting,
    build_presentation_frame,
    clean_report_columns,
    style_report_preview,
    write_ttl_average_formulas,
)


def build_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Country": "AU", "Category": "W&D", "Channel": "HN", "SKU": "H15 Pro Heat", "MAY": 0.84, "JUN": 1.0, "Trend": "▲"},
            {"Country": "AU", "Category": "W&D", "Channel": "HN", "SKU": "H16 Pro Steam", "MAY": 0.19, "JUN": 0.24, "Trend": "▲"},
            {"Country": "AU", "Category": "W&D", "Channel": "HN", "SKU": "TTL", "MAY": "", "JUN": "", "Trend": "▲"},
            {"Country": "AU", "Category": "W&D", "Channel": "JB", "SKU": "H15 Pro Heat", "MAY": 0.26, "JUN": 0.20, "Trend": "▼"},
            {"Country": "AU", "Category": "W&D", "Channel": "BSR", "SKU": "H16 Pro Steam", "MAY": "not visited", "JUN": "not visited", "Trend": ""},
            {"Country": "AU", "Category": "W&D", "Channel": "", "SKU": "TTL", "MAY": "", "JUN": "", "Trend": "▼"},
        ]
    )


def write_workbook(path: Path, frame: pd.DataFrame) -> None:
    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        frame.to_excel(writer, sheet_name="Report Preview", index=False)
        wb = writer.book
        pct = wb.add_format({"num_format": "0%"})
        green = wb.add_format({"bg_color": "#C6EFCE", "font_color": "#006100"})
        red = wb.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})
        cat_text = wb.add_format({"bg_color": "#0070C0", "font_color": "#FFFFFF", "bold": True})
        cat_pct = wb.add_format({"bg_color": "#0070C0", "font_color": "#FFFFFF", "bold": True, "num_format": "0%"})
        ch_text = wb.add_format({"bg_color": "#A6CAEC", "bold": True})
        ch_pct = wb.add_format({"bg_color": "#A6CAEC", "bold": True, "num_format": "0%"})
        ws = writer.sheets["Report Preview"]
        write_ttl_average_formulas(ws, frame, ["MAY", "JUN"], pct, cat_pct, ch_pct)
        apply_report_preview_formatting(ws, frame, "JUN", "MAY", "Trend", green, red, cat_text, ch_text)


class ReportFormattingTests(unittest.TestCase):
    def test_excel_colours_and_conditional_formats(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"
            write_workbook(path, build_frame())
            ws = load_workbook(path)["Report Preview"]

            # Category TTL row (position 5 -> excel row 7): strong blue fill, white.
            self.assertEqual(ws.cell(row=7, column=1).fill.fgColor.rgb, "FF0070C0")
            # Channel TTL row (position 2 -> excel row 4): lighter fill.
            self.assertEqual(ws.cell(row=4, column=1).fill.fgColor.rgb, "FFA6CAEC")
            # Detail row (position 0 -> excel row 2): no fill.
            self.assertIsNone(ws.cell(row=2, column=1).fill.patternType)
            # TTL JUN cell carries an AVERAGE formula and the blue fill.
            jun_cell = ws.cell(row=7, column=6)
            self.assertTrue(str(jun_cell.value).startswith("=IFERROR(AVERAGE"))
            self.assertEqual(jun_cell.fill.fgColor.rgb, "FF0070C0")
            # Conditional formatting present (value column + trend column).
            self.assertGreaterEqual(len(list(ws.conditional_formatting)), 2)

    def test_styler_marks_values_and_subtotals(self):
        frame = build_frame()
        rendered = style_report_preview(frame, "JUN").to_html()
        # Category TTL row shaded strong blue; green/red present for detail values.
        self.assertIn("#0070C0", rendered)
        self.assertIn("#C6EFCE", rendered)  # JUN 1.0 -> green
        self.assertIn("#FFC7CE", rendered)  # JUN 0.24 / 0.20 -> red

    def test_clean_report_columns_keeps_only_rolling_month_pair(self):
        frame = pd.DataFrame(
            [
                {
                    "Country": "AU",
                    "Category": "Robot",
                    "Channel": "JB",
                    "SKU": "X60 Ultra",
                    "APR": 0.1,
                    "MAY": 0.2,
                    "JUN": 0.3,
                    "JUL": 0.4,
                    "Trend": "▲",
                    "New": "",
                    "Note": "",
                }
            ]
        )

        cleaned = clean_report_columns(
            output_df=frame,
            month_label="JUL",
            previous_month_label="JUN",
            trend_column="Trend",
            keep_history_columns=False,
        )

        self.assertEqual(
            list(cleaned.columns),
            ["Country", "Category", "Channel", "SKU", "JUN", "JUL", "Trend", "New", "Note"],
        )

    def test_presentation_frame_copies_ttl_values_only(self):
        frame = build_frame()
        presentation = build_presentation_frame(frame, "MAY", "JUN", "Trend")

        self.assertEqual(
            list(presentation.columns),
            ["Country", "Category", "Channel", "MAY", "JUN", "Trend", "Key Points"],
        )
        self.assertEqual(len(presentation), 2)
        self.assertEqual(presentation["Key Points"].tolist(), ["", ""])
        self.assertNotIn("SKU", presentation.columns)
        self.assertEqual(presentation.loc[0, "Channel"], "HN")
        self.assertEqual(presentation.loc[0, "JUN"], "")
        self.assertEqual(presentation.loc[1, "Channel"], "")
        self.assertEqual(presentation.loc[1, "Trend"], "▼")

    def test_presentation_sheet_is_values_and_formatted(self):
        frame = pd.DataFrame(
            [
                {"Country": "AU", "Category": "W&D", "Channel": "HN", "MAY": 0.62, "JUN": 0.95, "Trend": "▲", "Key Points": ""},
                {"Country": "AU", "Category": "W&D", "Channel": "", "MAY": 0.41, "JUN": 0.53, "Trend": "▼", "Key Points": ""},
            ]
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "presentation.xlsx"
            with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
                frame.to_excel(writer, sheet_name="For Presentation", index=False)
                wb = writer.book
                green = wb.add_format({"bg_color": "#C6EFCE", "font_color": "#006100"})
                red = wb.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})
                cat_text = wb.add_format({"bg_color": "#0070C0", "font_color": "#FFFFFF", "bold": True})
                cat_pct = wb.add_format({"bg_color": "#0070C0", "font_color": "#FFFFFF", "bold": True, "num_format": "0%"})
                ch_text = wb.add_format({"bold": True})
                ch_pct = wb.add_format({"bold": True, "num_format": "0%"})
                apply_presentation_formatting(
                    writer.sheets["For Presentation"],
                    frame,
                    "MAY",
                    "JUN",
                    "Trend",
                    green,
                    red,
                    cat_text,
                    cat_pct,
                    ch_text,
                    ch_pct,
                )

            ws = load_workbook(path, data_only=False)["For Presentation"]
            self.assertIsNone(ws.cell(row=2, column=1).fill.patternType)
            self.assertEqual(ws.cell(row=3, column=1).fill.fgColor.rgb, "FF0070C0")
            self.assertEqual(ws.cell(row=3, column=7).value, None)
            self.assertFalse(str(ws.cell(row=2, column=5).value).startswith("="))
            self.assertGreaterEqual(len(list(ws.conditional_formatting)), 2)


if __name__ == "__main__":
    unittest.main()
