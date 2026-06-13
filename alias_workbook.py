from __future__ import annotations

import json
import shutil
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from identifier_matching import normalize_decision


ALIAS_SHEET_NAME = "_Alias Map"
ALIAS_PAYLOAD_VERSION = 1
CHUNK_SIZE = 30000


def alias_payload(decisions: Iterable[dict[str, object]]) -> dict[str, object]:
    return {
        "version": ALIAS_PAYLOAD_VERSION,
        "description": "Embedded identifier alias decisions for Calculator Online.",
        "decisions": [normalize_decision(item) for item in decisions],
    }


def _payload_text(decisions: Iterable[dict[str, object]]) -> str:
    return json.dumps(alias_payload(decisions), ensure_ascii=False, sort_keys=True)


def _payload_chunks(decisions: Iterable[dict[str, object]]) -> list[str]:
    text = _payload_text(decisions)
    return [text[index : index + CHUNK_SIZE] for index in range(0, len(text), CHUNK_SIZE)] or [""]


def normalize_alias_payload(data: object) -> list[dict[str, object]]:
    if isinstance(data, dict):
        data = data.get("decisions", [])
    if not isinstance(data, list):
        raise ValueError("Embedded alias map must be a list or {'decisions': [...]} payload.")
    return [normalize_decision(item) for item in data if isinstance(item, dict)]


def load_alias_decisions_from_workbook(path: Path | None) -> list[dict[str, object]]:
    if not path or not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if ALIAS_SHEET_NAME not in workbook.sheetnames:
            return []
        sheet = workbook[ALIAS_SHEET_NAME]
        chunks: list[str] = []
        for row in sheet.iter_rows(min_row=6, max_col=2, values_only=True):
            index, chunk = row
            if index in (None, "") and chunk in (None, ""):
                continue
            if chunk is not None:
                chunks.append(str(chunk))
        payload = "".join(chunks)
        if not payload:
            return []
        return normalize_alias_payload(json.loads(payload))
    finally:
        workbook.close()


def write_alias_sheet_openpyxl(workbook: Workbook, decisions: Iterable[dict[str, object]]) -> None:
    if ALIAS_SHEET_NAME in workbook.sheetnames:
        del workbook[ALIAS_SHEET_NAME]
    sheet = workbook.create_sheet(ALIAS_SHEET_NAME)
    sheet.sheet_state = "hidden"
    sheet["A1"] = "version"
    sheet["B1"] = ALIAS_PAYLOAD_VERSION
    sheet["A2"] = "payload_format"
    sheet["B2"] = "json_chunks"
    sheet["A3"] = "description"
    sheet["B3"] = "Embedded identifier alias decisions for Calculator Online."
    sheet["A5"] = "chunk_index"
    sheet["B5"] = "payload_chunk"
    for offset, chunk in enumerate(_payload_chunks(decisions), start=1):
        sheet.cell(row=5 + offset, column=1, value=offset)
        sheet.cell(row=5 + offset, column=2, value=chunk)


def write_alias_sheet_xlsxwriter(workbook, decisions: Iterable[dict[str, object]]) -> None:
    sheet = workbook.add_worksheet(ALIAS_SHEET_NAME)
    sheet.hide()
    sheet.write(0, 0, "version")
    sheet.write(0, 1, ALIAS_PAYLOAD_VERSION)
    sheet.write(1, 0, "payload_format")
    sheet.write(1, 1, "json_chunks")
    sheet.write(2, 0, "description")
    sheet.write(2, 1, "Embedded identifier alias decisions for Calculator Online.")
    sheet.write(4, 0, "chunk_index")
    sheet.write(4, 1, "payload_chunk")
    for offset, chunk in enumerate(_payload_chunks(decisions), start=1):
        sheet.write(4 + offset, 0, offset)
        sheet.write(4 + offset, 1, chunk)


def copy_workbook_with_alias_decisions(
    source: Path,
    target: Path,
    decisions: Iterable[dict[str, object]],
) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, target)
    workbook = load_workbook(target)
    try:
        write_alias_sheet_openpyxl(workbook, decisions)
        workbook.save(target)
    finally:
        workbook.close()
    return target


def create_alias_decisions_workbook(
    target: Path,
    decisions: Iterable[dict[str, object]],
) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Alias Carrier"
    workbook.active["A1"] = "Calculator Online alias carrier"
    write_alias_sheet_openpyxl(workbook, decisions)
    workbook.save(target)
    workbook.close()
    return target
